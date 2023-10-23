# -*- coding: utf-8 -*-
# ======================================
# 電子マネー管理システム
# MariaデータベースからデータExcel及びPDF出力
# 内容編集/出力モジュール
# [環境]20
#   Python 3.10.8
#   VSCode 1.64
#   <拡張>
#     |- Python  V2021.12
#     |- Pylance V2021.12
#
# [更新履歴]
#   2023/2/19  新規作成
# ======================================
from datetime import datetime
import datetime
import jpholiday
import os
import win32com.client
import pandas as pd
import openpyxl 
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from unmerge import excel_operate
from emoneydbclass import DataBaseClass

####
# 初期処理
####
class dbReportEdit:
    def __init__(self,parm,path,syear,smonth,sday,eyear,emonth,eday,prec,block):  
        
        # クラス初期化               
        self.resdb = DataBaseClass(parm)
        self.file_out_path  = path  
        self.SYEAR = syear
        self.SMONTH = smonth
        self.SDAY = sday
        self.EYEAR = eyear
        self.EMONTH = emonth
        self.EDAY = eday 
        self.prec = prec
        self.block = block
    ########################
    # カード種類別集計表
    ########################
    def print_syubetsu(self,df_card,df_paylog1):
        # 決済種別を結合
        df_paylog = pd.merge(df_paylog1,df_card, left_on='paycardcd', right_on='cardcode') 
        # 決済種別毎に集計
        df_sum_card = df_paylog[['cardname','payprice']].groupby('cardname').sum()
        
        #データをExcelに出力    
        with pd.ExcelWriter(f'{self.file_out_path}') as writer:
            df_sum_card.to_excel(writer,startrow=3,startcol=1,sheet_name='決済種別')  

        #Excelシートの加工
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb.worksheets[0]

        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A4
        # 印刷の向きを設定
        wps.orientation = sh.ORIENTATION_PORTRAIT

        sh.cell(row=1, column=2).value='決済種別売上集計表'
        #sh.cell(row=1, column=3).value='（電子マネー）'

        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2

        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
        
        adjusted_width = (max_length + 1) * 1.0   
        sh.column_dimensions[col[0].column_letter].width = adjusted_width
        
        #表記を修正
        sh.cell(row=4, column=2).value='決済種別'
        sh.cell(row=4, column=3).value='決済金額'    

        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 30 #明細種別名称
        sh.column_dimensions['C'].width = 30 #決済金額

        #決済金額合計をセット
        # 最終行の取得
        maxr = sh.max_row

        # 計算用の変数goukeiを定義
        goukei = 0

        # 所定範囲の数値を計算
        for i in range(5, maxr+1):
            kingaku = sh['C' + str(i)].value
            goukei += int(kingaku)

        # 合計値をセルに出力 中央ぞろえ
        sh['C' + str(maxr+1)].value = goukei
        sh['B' + str(maxr+1)].value = "合　計"
        font = Font(bold=True)
        sh['B' + str(maxr+1)].font = font
        sh['B' + str(maxr+1)].alignment = Alignment(horizontal="centerContinuous")

        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        for row_num in range(4,maxr+2):
            for col_num in range(2,4):
                sh.cell(row=row_num ,column=col_num).border = border

        #決済金額のフォーマットを変更
        for i in range(5,maxr+2):
            sh.cell(row=i,column=3).number_format = "¥#,##0" 
        
        wb.save(f'{self.file_out_path}')
        
        return 0
        
    ######################
    # 設置場所別集計表出力
    ######################
    def print_place(self,df_paylog,sheet_name):
        
        #データをExcelに出力    
        with pd.ExcelWriter(f'{self.file_out_path}', mode='a') as writer:
            df_paylog.to_excel(writer,startrow=3,startcol=1,sheet_name=sheet_name) 

        #Excelシートの加工開始
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb[sheet_name]

        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A4
        # 印刷の向きを設定
        wps.orientation = sh.ORIENTATION_PORTRAIT

        sh.cell(row=1, column=2).value='設置場所別売上集計表'
        sh.cell(row=1, column=3).value=sheet_name

        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
            
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2
        
        #表記を修正
        sh.cell(row=4, column=2).value='設置場所名'
        sh.cell(row=4, column=3).value='決済金額'  

        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            #column = col[0].column

            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
            
            adjusted_width = (max_length + 1) * 1.0   
            sh.column_dimensions[col[0].column_letter].width = adjusted_width

        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 40 #設置場所
        sh.column_dimensions['C'].width = 30 #決済金額

        #決済金額合計をセット
        # 最終行の取得
        maxr = sh.max_row
        
        # 計算用の変数goukeiを準備
        goukei = 0
        
        # 所定範囲の数値を計算
        for i in range(5, maxr+1):
            kingaku = sh['C' + str(i)].value
            goukei += kingaku
        
        # 合計値をセルに出力 中央ぞろえ
        sh['C' + str(maxr+1)].value = goukei
        sh['B' + str(maxr+1)].value = "合　計"
        font = Font(bold=True)
        sh['B' + str(maxr+1)].font = font
        sh['B' + str(maxr+1)].alignment = Alignment(horizontal="centerContinuous")
            
        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)

        for row_num in range(4,maxr+2):
            for col_num in range(2,4):
                sh.cell(row=row_num ,column=col_num).border = border

        #決済金額のフォーマットを変更
        for i in range(5,maxr+2):
            sh.cell(row=i,column=3).number_format = "¥#,##0"
            
        # save xlsx file
        wb.save(f'{self.file_out_path}')
        
        return 0
    ##########################
    # 金種別選択回数集計表出力
    # ########################  
    def print_kinsyu(self,df_paylog1,df_paylog2,sheet_name):
        #データをExcelに出力
        sheet_name2 = sheet_name + '_2'
        with pd.ExcelWriter(f'{self.file_out_path}', mode='a') as writer:
            df_paylog1.to_excel(writer,startrow=3,startcol=1,sheet_name=sheet_name)  
            df_paylog2.to_excel(writer,startrow=1,startcol=1,sheet_name=sheet_name2) 

        # #日付欄の結合を外す
        Inputfile = f'{self.file_out_path}'
        Excel = excel_operate(Inputfile)
        Excel.get_merged_cells_location()
        Excel.break_merged_cells() 

        #Excelシートの加工
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb[sheet_name]
        shx = wb[sheet_name2]
        
        #表記を修正
        sh.cell(row=4, column=4).value='決済時刻'
        sh.cell(row=5, column=3).value=''  
        #sh.cell(row=6, column=3).value='決済種別'  
        sh.cell(row=6, column=3).value='決済金額'      
        
        # 最終行の取得
        maxr = sh.max_row
        maxc = sh.max_column
        maxr2 = shx.max_row+1
        #一部結合中央揃え    
        sh.merge_cells(start_row=4,start_column=4,end_row=4,end_column=maxc) 
                    
        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A3
        # 印刷の向きを設定
        wps.orientation = sh.ORIENTATION_LANDSCAPE
            
        sh.cell(row=1, column=2).value='金種別利用回数集計表'
        sh.cell(row=1, column=3).value=sheet_name
            
        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2
            
        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
                
            adjusted_width = (max_length + 1) * 1.3   
            sh.column_dimensions[col[0].column_letter].width = adjusted_width
                
        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 25 #売上日
        #sh.column_dimensions['C'].width = 15 #決済種別
        sh.column_dimensions['C'].width = 15 #決済金額    
        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
            
        for row_num in range(4,maxr+1):    
            for col_num in range(3,maxc+1):
                sh.cell(row=row_num ,column=col_num).border = border
            
            
        #決済金額の表示フォーマットを変更
        for i in range(7,maxr+1):
            for j in range(2,maxc+1):
                #金種
                sh.cell(row=i,column=j).number_format = "#,##0"
        for i in range(7,maxr+1):
            sh.cell(row=i,column=3).number_format = "#,##0"

        sh.cell(row=maxr,column=2).value = "合計"
        sh.cell(row=5,column=maxc).value = "合計"
        sh.cell(row=6,column=2).value = "決済日"
            
        #金種別の合計表示
        ft = Font(bold=True)
        sh.cell(row=maxr+2,column=2).value = "金種別件数合計"
        sh.cell(row=maxr+2,column=2).font = ft
        #for i in range(5,12):
        for i in range(5,maxr2):
        #金種
            copy = shx.cell(row = i, column =2).value
            sh.cell(row = maxr+i-2, column = 3 , value = copy) .number_format = "#,##0"
            sh.cell(row = maxr+i-2, column = 3 , value = copy).font = ft
            sh.cell(row = maxr+i-2, column = 3 , value = copy).alignment = Alignment(horizontal = 'center', 
                                                                                            vertical = 'center')
        #件数合計
        for v in range(3,41):
            if shx.cell(row = 3, column = v).value == "Total": #合計行検索
                for p in range(4,maxr2):
                    copy2 = shx.cell(row = p, column = v).value
                    sh.cell(row = maxr+p-2, column = maxc , value = copy2).number_format = "#,##0"
                break
                
        #祝祭日の背景に色をつける
        # 会社特有の休日
        #company_holiday = ['2018-01-02','2018-01-03','2018-12-28','2018-12-31']        
        ymdx = 99999999
        for row_num in range(7,maxr):
            ymd = sh.cell(row=row_num,column=2).value
            wmd = str(ymd)
            if ymd != ymdx and ymd != None:
                #sh.unmerge_cells(row=row_num,column=2)
                ymdx = sh.cell(row=row_num,column=2).value   
                y = wmd[0:4]
                m = wmd[4:6]
                d = wmd[6:8]
                sh.cell(row=row_num,column=2).number_format = "###0"
                dt = datetime.date(int(y),int(m),int(d))
                cel = sh.cell(row=row_num,column=2)

                #土日、祝祭日判定してセルに色をつける   
                # 通常の土日
                if dt.weekday() == 5:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ffb76e')
            
                if dt.weekday() == 6:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ff2d3d')

                # 祝日
                dy = int(y)
                dm = int(m)
                dd = int(d)
                #祝日判定
                res_horiday = jpholiday.is_holiday_name(datetime.date(dy,dm,dd))
                if res_horiday != None:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='8eef6e')
                    
        #ワーク用シートの削除とブックの保存
        #wb.remove(wb.worksheets[3])
        wb.remove(wb[sheet_name2])
        wb.save(f'{self.file_out_path}')
        
        return 0
    
    ##########################
    # 金種別選択回数集計表出力
    # 新バージョン 2023/9/22
    # ########################  
    def print_kinsyu2(self,data_kinsyu,kinsyu_sum_data,sheet_name):
        # sheetの作成
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        #sh = wb[sheet_name]
        if sheet_name == '金種別(現金)':
            wb.create_sheet(index=3, title=sheet_name)
        else:
            wb.create_sheet(index=4, title=sheet_name) 
        sh = wb.get_sheet_by_name(sheet_name) 
        
        #ヘッダー部のセット
        #1行目
        row_no = 1
        sh.cell(row=row_no, column=2).value='金種別利用回数集計表'
        sh.cell(row=row_no, column=3).value=sheet_name        
        #2行目
        row_no = 2
        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')  
        sh.cell(row=row_no, column=2).value=str1
        sh.cell(row=row_no, column=3).value=str2
        #4行目
        row_no = 4
        sh.cell(row=row_no, column=4).value='決済時刻'
        #5行目
        row_no = 5
        col_no = 4        
        for x, row_data in enumerate(data_kinsyu):
            for y, cell_data in enumerate(row_data):
                if x == 0 and y > 1:
                    sh.cell(row=row_no,column=col_no).value=cell_data
                    col_no += 1
        sh.cell(row=row_no,column=col_no).value='合計'    
        #6行目
        row_no = 6
        sh.cell(row=row_no, column=2).value='決済日'  
        sh.cell(row=row_no, column=3).value='決済金額'
        
        #明細部のセット
        #7行目
        row_no = 7
        col_no = 2
        no_sum = 0
        kinsyu_item_count = len(data_kinsyu) - 1
        for x, row_data in enumerate(data_kinsyu):
            row_data_count = len(row_data)
            for y, cell_data in enumerate(row_data):
                if x > 0:
                    #明細を順次セット
                    sh.cell(row=row_no,column=col_no).value=cell_data
                    if cell_data != '' and y > 1:  
                        no_sum = no_sum + int(cell_data)
                    col_no += 1
            #利用回数総合計を表示
            if x > 0:
                sh.cell(row=row_no,column=col_no+1).value=no_sum
                #明細行最後の処理
                if x == kinsyu_item_count:
                    sh.cell(row=row_no,column=2).value='合計'
                    sh.cell(row=row_no,column=3).value=''
                    sh.cell(row=row_no,column=col_no+1).value=no_sum            
                row_no += 1
                col_no = 2
                no_sum = 0
        
        # 最終行の取得
        maxr = sh.max_row
        maxc = sh.max_column
        #決済日を日ごとにマージする  
        mer_str_no = 7  
        mer_date = sh.cell(row=7,column=2).value
        for i in range(7,maxr+1):
            if mer_date != sh.cell(row=i,column=2).value and  sh.cell(row=i,column=2).value != '':
                sh.merge_cells(start_row=mer_str_no,start_column=2,end_row=i-1,end_column=2)
                mer_date = sh.cell(row=i,column=2).value 
                mer_str_no = i               
                    
        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A3
        # 印刷の向きを設定
        wps.orientation = sh.ORIENTATION_LANDSCAPE
            
        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
                
            adjusted_width = (max_length + 1) * 1.3   
            sh.column_dimensions[col[0].column_letter].width = adjusted_width
                
        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 25 #売上日
        sh.column_dimensions['C'].width = 15 #決済金額    
        #罫線引く
        sh.merge_cells(start_row=4,start_column=2,end_row=5,end_column=3)
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
            
        for row_num in range(4,maxr+1):    
            for col_num in range(2,maxc+1):
                sh.cell(row=row_num ,column=col_num).border = border
                
        #決済金額の表示フォーマットを変更
        for i in range(7,maxr+1):
            for j in range(3,maxc+1):
                #金種
                sh.cell(row=i,column=j).number_format = "#,##0"
        #for i in range(7,maxr+1):
        #    sh.cell(row=i,column=3).number_format = "#,##0"
            
        #金種別の合計表示
        ft = Font(bold=True)
        sh.cell(row=maxr+2,column=2).value = "金種別件数合計"
        sh.cell(row=maxr+2,column=2).font = ft
        
        sum_count = len(kinsyu_sum_data)
        kinsyu_sum_tuple_key = list(kinsyu_sum_data.keys())
        kinsyu_sum_tuple_value = list(kinsyu_sum_data.values())
        idx = 0
        kinsyu_sum_tuple_value_sum = 0
        
        for i in range(maxr+3,maxr+3+sum_count):
            #金種をセット
            sh.cell(row = i, column = 3 , value = kinsyu_sum_tuple_key[idx]).number_format = "#,##0"
            sh.cell(row = i, column = 3 , value = kinsyu_sum_tuple_key[idx]).font = ft
            #数値をセット
            kinsyu_sum_tuple_value_sum += int(kinsyu_sum_tuple_value[idx]) 
            sh.cell(row = i, column = maxc , value = kinsyu_sum_tuple_value[idx]).alignment = Alignment(horizontal = 'center', 
                                                                                                vertical = 'center') 
            idx += 1 
        #総合計値をセット
        sh.cell(row = i+1, column = 3 , value = '総合計')
        sh.cell(row = i+1, column = maxc, value = kinsyu_sum_tuple_value_sum).number_format = "#,##0"
                
        """ #祝祭日の背景に色をつける
        # 会社特有の休日
        #company_holiday = ['2018-01-02','2018-01-03','2018-12-28','2018-12-31']        
        ymdx = 99999999
        for row_num in range(7,maxr):
            ymd = sh.cell(row=row_num,column=2).value
            if ymd != '':
                wmd = str(ymd)
                if ymd != ymdx and ymd != None:
                    #sh.unmerge_cells(row=row_num,column=2)
                    ymdx = sh.cell(row=row_num,column=2).value   
                    y = wmd[0:4]
                    m = wmd[4:6]
                    d = wmd[6:8]
                    sh.cell(row=row_num,column=2).number_format = "###0"
                    dt = datetime.date(int(y),int(m),int(d))
                    cel = sh.cell(row=row_num,column=2)

                    #土日、祝祭日判定してセルに色をつける   
                    # 通常の土日
                    if dt.weekday() == 5:
                        sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ffb76e')
            
                    if dt.weekday() == 6:
                        sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ff2d3d')

                    # 祝日
                    dy = int(y)
                    dm = int(m)
                    dd = int(d)
                    #祝日判定
                    res_horiday = jpholiday.is_holiday_name(datetime.date(dy,dm,dd))
                    if res_horiday != None:
                        sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='8eef6e') """
                    
        #ワーク用シートの削除とブックの保存
        wb.save(f'{self.file_out_path}')
        
        return 0
    
    ####################
    # 時間別集計表出力
    ####################
    def print_jikan(self,df_paylog1,sheet_name):        
            
        #データをExcelに出力     
        with pd.ExcelWriter(f'{self.file_out_path}', mode='a') as writer:
            df_paylog1.to_excel(writer,startrow=3,startcol=1,sheet_name=sheet_name)   
            
        #Excelシートの加工
        wb = openpyxl.load_workbook(f'{self.file_out_path}')
        sh = wb[sheet_name]
        
        #天気情報挿入
        sh.insert_cols(3, 4) #列挿入
        for i in range(7,len(df_paylog1)+6):
            ymdt = sh.cell(row=i, column=2).value
            wmdt = str(ymdt) 
            wdf = self.resdb.weather_get(wmdt[0:4],wmdt[4:6],wmdt[6:8],self.prec,self.block) # 日付で天気検索
            sh.cell(row=i, column=3).value = wdf[0][1] #天気（昼）
            sh.cell(row=i, column=4).value = wdf[0][2] #天気（夜）
            sh.cell(row=i, column=5).value = wdf[0][3] #最高気温
            #sh.cell(row=i,column=5).number_format = "0.00"
            sh.cell(row=i, column=6).value = wdf[0][4] #最低気温
            #sh.cell(row=i,column=6).number_format = "0.00"
                    
        #用紙設定
        wps = sh.page_setup
        # 用紙サイズを設定
        wps.paperSize = sh.PAPERSIZE_A3
        # 印刷の向きを設定wq
        wps.orientation = sh.ORIENTATION_LANDSCAPE
        wps.fitToWidth = 1
        wps.fitToHeight = 1
        sh.sheet_properties.pageSetUpPr.fitToPage = True       
        #ヘッダーセット
        sh.cell(row=1, column=2).value='売上日・時間別売上集計表'
        sh.cell(row=1, column=3).value=sheet_name
            
        str1 = (f'{self.SYEAR} 年 {self.SMONTH} 月 {self.SDAY} 日  ～')
        str2 = (f'{self.EYEAR} 年 {self.EMONTH} 月 {self.EDAY} 日')
        sh.cell(row=2, column=2).value=str1
        sh.cell(row=2, column=3).value=str2
                
        # 決済金額合計をセット
        # 最終行の取得
        maxr = sh.max_row
        maxc = sh.max_column
            
        #罫線引く
        side = Side(style='thin', color='000000')
        border = Border(top=side, bottom=side, left=side, right=side)
            
        for row_num in range(4,maxr+1):    
            for col_num in range(2,maxc+1):
                sh.cell(row=row_num ,column=col_num).border = border
            
            
        #決済金額のフォーマットを変更
        for i in range(5,maxr+1):
            for j in range(4,maxc+1):
                sh.cell(row=i,column=j).number_format = "#,##0"
            
        sh.cell(row=maxr,column=2).value = "合計"
        sh.cell(row=5,column=maxc).value = "合計"
        sh.cell(row=6,column=2).value = "売上日"
        sh.cell(row=6,column=3).value = "天気（6:00～18:00)"
        sh.cell(row=6,column=4).value = "天気（18:00～翌6:00)"
        sh.cell(row=6,column=5).value = "最高気温"
        sh.cell(row=6,column=6).value = "最低気温"
            
        # セル幅を自動調整
        for col in sh.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))        
                
            adjusted_width = (max_length + 1) * 1.3   
            sh.column_dimensions[col[0].column_letter].width = adjusted_width
                
        #部分的にセル幅を修正
        sh.column_dimensions['B'].width = 25 #売上日
        sh.column_dimensions['C'].width = 25 #天気（昼）
        sh.column_dimensions['D'].width = 25 #天気（夜）
        sh.column_dimensions['E'].width = 10 #最高気温
        sh.column_dimensions['F'].width = 10 #最低気温
        sh.column_dimensions['G'].width = 7
        #表記を修正
        sh.cell(row=5, column=2).value=' '  
        sh.cell(row=5, column=6).value='時間　→' 
                    
        #祝祭日の背景に色をつける
        # 会社特有の休日
        #company_holiday = ['2018-01-02','2018-01-03','2018-12-28','2018-12-31']
        ymdx = 99999999
        for row_num in range(7,maxr):
            ymd = sh.cell(row=row_num,column=2).value
            wmd = str(ymd)
            if ymd != ymdx and ymd != None:
                ymdx = sh.cell(row=row_num,column=2).value   
                y = wmd[0:4]
                m = wmd[4:6]
                d = wmd[6:8]
                sh.cell(row=row_num,column=2).number_format = "###0"
                dt = datetime.date(int(y),int(m),int(d))
                cel = sh.cell(row=row_num,column=2)           

                #土日、祝祭日判定してセルに色をつける   
                # 通常の土日
                if dt.weekday() == 5:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ffb76e')
                
                if dt.weekday() == 6:
                    sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='ff2d3d')

            # 祝日
            dy = int(y)
            dm = int(m)
            dd = int(d)
            #祝日判定
            res_horiday = jpholiday.is_holiday_name(datetime.date(dy,dm,dd))
            if res_horiday != None:
                sh[cel.coordinate].fill = PatternFill(patternType='solid', fgColor='8eef6e')
        
            # #会社の休日
            # if date.strftime("%Y-%m-%d") in company_holiday:
            #     return Tru                
        
        # save xlsx file
        wb.save(f'{self.file_out_path}')
        
        return 0
    #
    # 時間別集計表出力（テストトライアル）
    #
    def print_jikan2(self,df_paylog1,sheet_name):        
            
        #データをExcelに出力     
        with pd.ExcelWriter(f'{self.file_out_path}', mode='a') as writer:
            df_paylog1.to_excel(writer,startrow=3,startcol=1,sheet_name=sheet_name) 
        
        print('テスト完了')  

    #
    # 出力されたEXCELシートをwin32apiを使ってPDF変換する
    #     
    def pdfconv(self,dir_out_filepath):
        excel = win32com.client.Dispatch("Excel.Application")
        # pdfへの変換
        path = dir_out_filepath + '/'
        wb = excel.Workbooks.Open(self.file_out_path)
        excelSheets = []
        for sheet in wb.Worksheets:
            excelSheets.append(sheet.name)
        #for i in range(1,5):
        sheet_num = len(excelSheets)
        for i in range(1,sheet_num+1):   
            wb.WorkSheets(i).Select()
            try:
                if os.path.exists(path + excelSheets[i-1] + '.pdf' ):
                    os.remove(path + excelSheets[i-1] + '.pdf' )
                wb.ActiveSheet.ExportAsFixedFormat(0, path + excelSheets[i-1] + '.pdf' )
            except:
                print('PDFファイルが正常に保存できませんでした')
                wb.Close()
                excel.Quit() 
            
        wb.Close()
        excel.Quit() 
    
    ###############################################################
    # ディストラクタ
    ###############################################################
    def __del__(self):
        #print('ディストラクタ呼び出し') 
        pass 


   
           
    